from openpyxl import Workbook
from openpyxl import load_workbook

def merging_two_files():
    wb = Workbook()
    sheet = wb.active

    month = input('Enter the numeric value of the month (format 01 - 12): ')
    year = input('Enter the numeric value of the year (format 2022): ')

    def append_rows(dst, dst_h, frm_ws, frm_h):
        row_index = 0
        for row in frm_ws.values:
            if row_index == 0:
                row_index += 1
                continue
            column = 0
            row_value = [None]*len(dst_h)
            for value in row:
                value_list = frm_h[column]
                try:
                    index_ws3 = dst_h.index(value_list)
                    row_value[index_ws3] = value
                except:
                    pass
                column += 1
            dst.append(row_value)

    len_list_import = 0
    while True:
        try:
            path_to_read_file1 = input('Input path to read import file: ')
            wb1 = load_workbook(path_to_read_file1)
            ws1 = wb1.active
            cell_range1 = ws1[1]
            
            for _ in ws1: len_list_import += 1  
            
            list_ws1 = []

            for cell in cell_range1:
                if cell.value == 'ЄДРПОУ отримувача': cell.value = 'ua_customer_code'
                elif cell.value == 'Назва отримувача': cell.value = 'c_importer_1'
                elif cell.value == 'Країна відправлення': cell.value = 'c_country_of_shipping'
                elif cell.value == 'Торгуюча країна': cell.value = 'c_trading_country'
                elif cell.value == 'Країна походження': cell.value = 'c_country_of_origin'
                elif cell.value == 'Іноземний відправник товару': cell.value = 'c_exporter'
                elif cell.value == 'Іноземний контрактоутримувач (зворотний бік ВМД)': cell.value = 'c_exporter_1'
                elif cell.value == 'Назва українського контрактоутримувача': cell.value = 'c_importer_2'
                elif cell.value == 'Митниця оформлення': cell.value = 'c_custom'
                elif cell.value == 'Базис постачання': cell.value = 'incoterms'
                elif cell.value == 'українського контрактоутримувача': cell.value = 'ua_customer_code_2'
                elif cell.value == 'Транспорт на кордоні': cell.value = 'c_transport_on_border'
                elif cell.value == 'Транспорт в Україні': cell.value = 'с_transport_in_UA'
                elif cell.value == 'Одиниця виміру': cell.value = 'c_cargo units'
                elif cell.value == 'Кількість у додаткових одиницях виміру': cell.value = 'c_cargo_count_units'
                elif cell.value == 'Вага нетто, кг': cell.value = 'c_gross_weight'
                elif cell.value == 'Фактурна вартість, USD': cell.value = 'c_invoice_amount'
                elif cell.value == 'Митна вартість, USD': cell.value = 'c_custom_value_amount'
                elif cell.value == 'Тип декларації': cell.value = 'c_md_type'
                elif cell.value == 'Режим': cell.value = 'c_md_type'
                elif cell.value == 'Код митного органу': cell.value = 'custom_office_code'
                elif cell.value == '№ ВМД': cell.value = 'c_declaration_number'
                elif cell.value == ' № товару': cell.value = 'c_cargo_number'
                elif cell.value == 'Код товару УКТЗЕД': cell.value = 'c_code_of_cargo'
                elif cell.value == 'Опис товару по УКТЗЕД': cell.value = 'c_cargo'
                else:cell.value = None
                list_ws1.append(cell.value)

            update_list = ['imp/exp', 'year', 'month']
            for i in update_list:
                list_ws1.append(i)
            break
                
        except:
            print('No such file or directory', '\n')
            choise = input('If you want to exit press the letter e or E: ')
            if choise == 'e' or choise == 'E':
                exit()
            else:
                continue

    while True:
        try:
            path_to_read_file2 = input('Input path to read export file: ')
            wb2 = load_workbook(path_to_read_file2)
            ws2 = wb2.active
            cell_range2 = ws2[1]

            list_ws2 = []
            
            for cell in cell_range2:
                if cell.value == 'ЄДРПОУ відправника': cell.value = 'ua_customer_code'
                elif cell.value == 'Назва відправника': cell.value = 'c_exporter'
                elif cell.value == 'Країна призначення': cell.value = 'c_destin_country'
                elif cell.value == 'Іноземний одержувач/контрактоутримувач': cell.value = 'c_importer_1'
                elif cell.value == 'Іноземний контрактоутримувач (зворотний бік ВМД)': cell.value = 'c_importer_2'
                elif cell.value == 'Назва українського контрактоутримувача': cell.value = 'c_exporter_1'
                elif cell.value == 'Митниця оформлення': cell.value = 'c_custom'
                elif cell.value == 'Базис постачання': cell.value = 'incoterms'
                elif cell.value == 'ЄДРПОУ українського контрактоутримувача': cell.value = 'ua_customer_code_2'
                elif cell.value == 'Транспорт на кордоні': cell.value = 'c_transport_on_border'
                elif cell.value == 'Транспорт в Україні': cell.value = 'с_transport_in_UA'
                elif cell.value == 'Одиниця виміру': cell.value = 'c_cargo units'
                elif cell.value == 'Кількість у додаткових одиницях виміру': cell.value = 'c_cargo_count_units'
                elif cell.value == 'Вага нетто, кг': cell.value = 'c_gross_weight'
                elif cell.value == 'Фактурна вартість, USD': cell.value = 'c_invoice_amount'
                elif cell.value == 'Митна вартість, USD': cell.value = 'c_custom_value_amount'
                elif cell.value == 'Тип декларації': cell.value = 'c_md_type'
                elif cell.value == 'Режим': cell.value = 'c_md_type'
                elif cell.value == 'Код митного органу': cell.value = 'custom_office_code'
                elif cell.value == '№ ВМД': cell.value = 'c_declaration_number'
                elif cell.value == '№ товару': cell.value = 'c_cargo_number'
                elif cell.value == 'Код товару УКТЗЕД': cell.value = 'c_code_of_cargo'
                elif cell.value == 'Опис товару по УКТЗЕД': cell.value = 'c_cargo'
                else:cell.value = None
                list_ws2.append(cell.value)

            break

        except:
            print('No such file or directory', '\n')
            choise = input('If you want to exit press the letter e or E: ')
            if choise == 'e' or choise == 'E':
                exit()
            else:
                continue
            
    list_ws3 = []

    for value in list_ws1:
        if value != None and value not in list_ws3:
            list_ws3.append(value)

    for value in list_ws2:
        if value != None and value not in list_ws3:
            list_ws3.append(value)

    rows_value = [list_ws3]

    append_rows(rows_value, list_ws3, ws1, list_ws1)
    append_rows(rows_value, list_ws3, ws2, list_ws2)


    index_cargo = 0
    index_imp_exp = 0
    index_year = 0
    index_month = 0

    for column_value in range(len(list_ws3)):
        if list_ws3[column_value] == 'c_cargo':
            index_cargo = column_value
        
        elif list_ws3[column_value] == 'imp/exp':
            index_imp_exp = column_value
        
        elif list_ws3[column_value] == 'year':
            index_year = column_value
        
        elif list_ws3[column_value] == 'month':
            index_month = column_value

    for i  in range(1, len(rows_value)): 
        try:  
            if len(rows_value[i][index_cargo]) > 80:
                rows_value[i][index_cargo] = rows_value[i][index_cargo][0:80]
        except:
            pass

        if not rows_value[i][index_imp_exp]: 
            if i < len_list_import:
                rows_value[i][index_imp_exp] = 'imp'
            else:
                rows_value[i][index_imp_exp] = 'exp'

        if not rows_value[i][index_year]: 
            rows_value[i][index_year] = f'{year}'

        if not rows_value[i][index_month]: 
            rows_value[i][index_month] = f'{month}'

    for i in range(len(rows_value)):
        for j in range(len(rows_value[i])):
            sheet.cell(row=i+1, column=j+1).value = rows_value[i][j]

    wb.save(f'megrating_file/{year}_{month}_imp_exp_db_UA.xlsx')

    print('Files processed successfully!')
